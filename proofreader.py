#!/usr/bin/env python3
"""Proofreader tool: checks JPEG images for spelling, grammar, translation, and brand guideline issues using Claude's vision API."""

from __future__ import annotations

import argparse
import base64
import json
import mimetypes
import sys
import time
from pathlib import Path

import anthropic
import openpyxl
from PIL import Image
import io
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).resolve().parent

# Load environment variables from .env file
load_dotenv(SCRIPT_DIR / ".env")

REFERENCES_DIR = SCRIPT_DIR / "references"
EMEA_PATH = REFERENCES_DIR / "EMEA Messaging Matrix-2[95].xlsx"
GLOBAL_CTA_PATH = REFERENCES_DIR / "Global_CTA_Matrix_Original Movies  Series_MASTER_ Dropdown_8.1.25-2.xlsx"
BRAND_GUIDELINES_PATH = REFERENCES_DIR / "brand_guidelines.pdf"
BRAND_RULES_PATH = REFERENCES_DIR / "brand_rules.txt"
COMMON_MISTAKES_PATH = REFERENCES_DIR / "common_mistakes_examples.pdf"
MASTER_EXAMPLES_PATH = REFERENCES_DIR / "master_examples.pdf"

SYSTEM_PROMPT = """You are an expert proofreader and brand compliance reviewer. You will be given an image (product packaging, marketing material, etc.) and possibly brand guidelines documents and rules.

Your job:
1. **OCR**: Extract ALL visible text from the image.
2. **Spelling**: Check every word for spelling errors in each detected language.
3. **Grammar**: Carefully check for grammar mistakes in all text. Pay close attention to subject-verb agreement (e.g., "battery life last" should be "battery life lasts"), incorrect tense, missing articles, wrong prepositions, and sentence structure errors. Report every grammar issue you find — do not skip any.
4. **Translation**: If multiple languages are expected, verify translations are accurate and consistent between languages.
5. **Approved Messaging Compliance**: If reference translation/messaging data is provided, compare ALL text in the image against the approved translations. Flag any text that uses unapproved wording, wrong capitalization style, or deviates from the approved CTAs and messaging.
6. **Brand Rules Compliance**: If explicit brand rules are provided (colors, fonts, spacing, layout measurements), meticulously check EVERY requirement.
   - **MEASUREMENTS**: Actively measure and calculate proportions by comparing element dimensions to the overall image dimensions. For example, if the rule states "container must be 25% of width", measure the container width and total width, calculate the percentage, and determine if it meets the requirement. Report your measurements (e.g., "Container is 24.8% of total width - does not meet exact 25% requirement" or "Container is 25% of total width - complies with requirement"). DO NOT ask for verification - make a determination based on your measurements.
   - **CONTAINER LOGO SIZING (CRITICAL)**: ALWAYS check if the container logo (text/logo inside the blue container) is the correct size. This is a COMMON MISTAKE. Measure the logo size vs container size and compare to the required percentages. For landscape: stacked logos should be 60% of container width, single-line logos should be 70%. For portrait: logos should be 50% of container height. Container logos are frequently too small - check this carefully every time.
   - **CONTAINER LOGO POSITIONING (CRITICAL)**: ALWAYS check if the container logo is centered within the container. This is a COMMON MISTAKE. Check if there is equal space on the left/right and top/bottom of the logo. Container logos are frequently off-center (pushed left, right, up, or down) - check both horizontal and vertical centering carefully every time, especially in portrait images.
   - **COLORS**: Verify colors match exactly (hex codes if visible in the image).
   - **FONTS**: If master examples are provided, COMPARE the fonts in the test image to the fonts shown in the master examples. Look at character shapes, weights, proportions, and styling. Flag if the test image uses noticeably different fonts from the master examples. For CTAs specifically, check that the first half (before pipe) is bold and the second half (after pipe) is regular weight, matching the master examples.
   - **SPACING**: Validate spacing ratios and measurements.
   - **BREAKOUTS**: Carefully examine any elements breaking out from the key art into the border/container. Check if they touch or come too close to the container logo (logo inside the blue container at the top). Measure the distance visually and flag if the breakout appears to touch or encroach on the logo.
   - Be confident in your measurements and flag ANY deviations from the specified exact values in the brand rules.
7. **Visual/Brand Compliance**: If master examples are provided, use them as visual references to compare fonts, layouts, styling, and overall appearance. If brand guidelines PDF is provided, use it to understand correct layouts and compare against the rules.

Return your analysis as a JSON object with this exact structure:
{
  "extracted_text": "All text found in the image, preserving layout as much as possible",
  "languages_detected": ["English", "Spanish"],
  "element_identification": {
    "title_treatment": "The exact text you identified as the Title Treatment (movie/series name)",
    "signature": "The exact text you identified as the Signature (logo above title treatment)",
    "cta": "The exact text you identified as the CTA (call to action below title treatment)",
    "container_logo": "The exact text you identified as the Container Logo (logo inside blue container)"
  },
  "issues": [
    {
      "category": "spelling" | "grammar" | "translation" | "messaging_compliance" | "brand_compliance" | "visual",
      "severity": "error" | "warning" | "info",
      "location": "Description of where on the image (e.g., 'front panel, top-right')",
      "text": "The problematic text or element",
      "problem": "What is wrong",
      "suggestion": "Suggested fix"
    }
  ],
  "summary": "Brief overall assessment"
}

Return ONLY the JSON object, no other text. If there are no issues in a category, still return the JSON with an empty issues array."""


# Map common language names to spreadsheet identifiers
LANGUAGE_TO_EMEA_COL = {
    "english": "EN (English)", "german": "DE (German)", "french": "FR (French)",
    "italian": "IT (Italian)", "spanish": "ES (Spanish)", "dutch": "NL (Dutch)",
    "swedish": "SE (Swedish)", "norwegian": "NO (Norwegian)", "danish": "DK (Danish)",
    "finnish": "FI (Finnish)", "portuguese": "PT (Portuguese)", "polish": "PL (Polish)",
    "turkish": "TR (Turkish)", "romanian": "RO (Romanian)", "czech": "CZ (Czech)",
    "hungarian": "HU (Hungarian)", "greek": "GR (Greek)",
}

LANGUAGE_TO_CTA_SHEET = {
    "english": "en-US", "german": "de-DE", "french": "fr-FR",
    "italian": "it-IT", "spanish": "es-ES", "dutch": "nl-NL",
    "swedish": "sv-SE", "norwegian": "nb-NO", "danish": "da-DK",
    "finnish": "fi-FI", "portuguese": "pt-PT", "polish": "pl-PL",
    "turkish": "tr-TR", "romanian": "ro-RO", "czech": "cs-CZ",
    "hungarian": "hu-HU", "greek": "el-GR", "arabic": "ar-AE",
    "japanese": "ja-JP", "korean": "ko-KR",
}


def extract_emea_references(languages: list[str]) -> str:
    """Extract relevant columns from the EMEA Messaging Matrix for the given languages."""
    if not EMEA_PATH.exists():
        return ""

    wb = openpyxl.load_workbook(EMEA_PATH, read_only=True, data_only=True)
    ws = wb["EMEA Messaging Matrix"]

    # Find header row and column indices
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Always include English + requested languages
    lang_keys = {"english"} | {l.strip().lower() for l in languages}
    col_indices = {}
    for lang_key in lang_keys:
        col_name = LANGUAGE_TO_EMEA_COL.get(lang_key)
        if col_name:
            for idx, val in enumerate(header_row):
                if val and col_name in str(val):
                    col_indices[col_name] = idx
                    break

    if not col_indices:
        wb.close()
        return ""

    # Extract the SOP guidelines column (col 0) + language columns
    lines = ["## EMEA Messaging Matrix — Approved Translations\n"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        row_list = list(row)
        guideline = row_list[0] if row_list[0] else ""
        entries = {}
        for col_name, idx in col_indices.items():
            val = row_list[idx] if idx < len(row_list) and row_list[idx] else ""
            if val:
                entries[col_name] = str(val).strip()
        if entries:
            if guideline:
                lines.append(f"Context: {guideline}")
            for col_name, val in entries.items():
                lines.append(f"  {col_name}: {val}")
            lines.append("")

    wb.close()
    return "\n".join(lines)


def extract_cta_references(languages: list[str]) -> str:
    """Extract relevant sheets from the Global CTA Matrix for the given languages."""
    if not GLOBAL_CTA_PATH.exists():
        return ""

    wb = openpyxl.load_workbook(GLOBAL_CTA_PATH, read_only=True, data_only=True)

    # Always include en-US + requested languages
    lang_keys = {"english"} | {l.strip().lower() for l in languages}
    sheets_to_read = set()
    for lang_key in lang_keys:
        sheet_name = LANGUAGE_TO_CTA_SHEET.get(lang_key)
        if sheet_name and sheet_name in wb.sheetnames:
            sheets_to_read.add(sheet_name)

    if not sheets_to_read:
        wb.close()
        return ""

    lines = ["## Global CTA Matrix — Approved Messaging\n"]
    for sheet_name in sorted(sheets_to_read):
        ws = wb[sheet_name]
        lines.append(f"### Locale: {sheet_name}\n")
        for row in ws.iter_rows(min_row=14, values_only=True):
            row_list = list(row)
            # Columns: context (B=1), messaging EN (C=2), messaging local (D=3),
            # title case (E=4), sentence case (F=5), upper case (G=6)
            context = row_list[1] if len(row_list) > 1 and row_list[1] else ""
            msg_en = row_list[2] if len(row_list) > 2 and row_list[2] else ""
            msg_local = row_list[3] if len(row_list) > 3 and row_list[3] else ""
            title_case = row_list[4] if len(row_list) > 4 and row_list[4] else ""
            sentence_case = row_list[5] if len(row_list) > 5 and row_list[5] else ""
            upper_case = row_list[6] if len(row_list) > 6 and row_list[6] else ""

            if any([msg_en, msg_local, title_case, sentence_case, upper_case]):
                parts = []
                if context:
                    parts.append(f"Context: {str(context).strip()}")
                if msg_en:
                    parts.append(f"  EN: {str(msg_en).strip()}")
                if msg_local:
                    parts.append(f"  Local: {str(msg_local).strip()}")
                if title_case:
                    parts.append(f"  Title Case: {str(title_case).strip()}")
                if sentence_case:
                    parts.append(f"  Sentence Case: {str(sentence_case).strip()}")
                if upper_case:
                    parts.append(f"  Upper Case: {str(upper_case).strip()}")
                lines.append("\n".join(parts))
                lines.append("")

    wb.close()
    return "\n".join(lines)


def load_reference_data(languages: list[str]) -> str:
    """Load all reference data for the given languages."""
    parts = []

    emea = extract_emea_references(languages)
    if emea:
        parts.append(emea)

    cta = extract_cta_references(languages)
    if cta:
        parts.append(cta)

    if not parts:
        return ""

    header = (
        "=== REFERENCE DATA: Approved Translations & Messaging ===\n"
        "The following are the officially approved translations and CTAs. "
        "Compare ALL text in the image against these references. "
        "Flag any text that does not match the approved translations — "
        "wrong wording, wrong capitalization, or unapproved messaging.\n\n"
    )
    return header + "\n\n".join(parts)


def compress_image(path: Path, max_dimension: int = 5000, max_size_mb: float = 4.0) -> bytes:
    """Compress an image to reduce file size while maintaining quality for analysis.

    Args:
        path: Path to the image file
        max_dimension: Maximum width or height in pixels
        max_size_mb: Maximum file size in MB (default 4.5MB to stay under 5MB API limit)
    """
    # Check original file size
    original_size_mb = path.stat().st_size / (1024 * 1024)

    # If file is already small (< 2MB), return original without compression
    if original_size_mb < 2.0:
        return path.read_bytes()

    img = Image.open(path)

    # Convert RGBA to RGB if necessary
    if img.mode == 'RGBA':
        background = Image.new('RGB', img.size, (255, 255, 255))
        background.paste(img, mask=img.split()[3])
        img = background
    elif img.mode != 'RGB':
        img = img.convert('RGB')

    # For very large files, start with more aggressive resizing
    width, height = img.size
    current_max_dimension = max_dimension

    # Adjust initial resize based on original file size
    if original_size_mb > 20:
        current_max_dimension = 2500  # Very aggressive for huge files
    elif original_size_mb > 15:
        current_max_dimension = 3000  # Aggressive for very large files
    elif original_size_mb > 10:
        current_max_dimension = 3500  # More aggressive for large files

    # Resize if image is too large
    if max(width, height) > current_max_dimension:
        ratio = current_max_dimension / max(width, height)
        new_size = (int(width * ratio), int(height * ratio))
        img = img.resize(new_size, Image.Resampling.LANCZOS)

    # Start with high quality and reduce if needed to stay under size limit
    quality = 98
    buffer = io.BytesIO()
    img.save(buffer, format='JPEG', quality=quality, optimize=True)
    compressed_size_mb = len(buffer.getvalue()) / (1024 * 1024)

    # If still too large, reduce quality iteratively (go as low as 60%)
    while compressed_size_mb > max_size_mb and quality > 60:
        quality -= 5
        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=quality, optimize=True)
        compressed_size_mb = len(buffer.getvalue()) / (1024 * 1024)

    # If STILL too large after quality reduction, resize further
    if compressed_size_mb > max_size_mb:
        current_width, current_height = img.size
        while compressed_size_mb > max_size_mb and max(current_width, current_height) > 800:
            # Reduce dimensions by 25% each iteration (more aggressive)
            current_width = int(current_width * 0.75)
            current_height = int(current_height * 0.75)
            img = img.resize((current_width, current_height), Image.Resampling.LANCZOS)

            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=quality, optimize=True)
            compressed_size_mb = len(buffer.getvalue()) / (1024 * 1024)

    return buffer.getvalue()


def encode_file(path: Path) -> tuple[str, str]:
    """Read a file and return (base64_data, media_type)."""
    mime_type, _ = mimetypes.guess_type(str(path))
    if mime_type is None:
        suffix = path.suffix.lower()
        mime_map = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
            ".pdf": "application/pdf",
        }
        mime_type = mime_map.get(suffix, "application/octet-stream")

    # Compress images to reduce request size
    if mime_type and mime_type.startswith("image/"):
        data = compress_image(path)
        mime_type = "image/jpeg"  # Always return JPEG after compression
    else:
        data = path.read_bytes()

    return base64.standard_b64encode(data).decode("utf-8"), mime_type


def build_content_blocks(image_path: Path, guidelines_path: Path | None, languages: str, asset_type: str = "General", is_prime_original: bool = False, pre_or_post: str = "Pre") -> list[dict]:
    """Build the content blocks for the Claude API request."""
    blocks = []

    is_pdf = image_path.suffix.lower() == ".pdf"

    # Detect orientation from image dimensions (skip for PDFs)
    if is_pdf:
        orientation = None
    else:
        img = Image.open(image_path)
        width, height = img.size
        orientation = "Landscape" if width > height else "Portrait"

    # Add the main file (PDF as document, image as image)
    file_data, file_mime = encode_file(image_path)
    if is_pdf:
        blocks.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": file_mime,
                "data": file_data,
            },
        })
    else:
        blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": file_mime,
                "data": file_data,
            },
        })

    # Determine color mode based on asset type
    color_mode = "CMYK" if asset_type in ["OOH", "T-Sides"] else "RGB"

    asset_info = f"Above is the {'document' if is_pdf else 'image'} to proofread. Expected languages: {languages}."

    # Add orientation instruction
    if orientation:
        asset_info += f"\n\nORIENTATION: {orientation}"
        if orientation == "Landscape":
            asset_info += "\nIMPORTANT: This is a LANDSCAPE image (wider than tall). Apply landscape-specific rules: Container at top, Border 2.5%, TT within 2.5% of border."
        else:
            asset_info += "\nIMPORTANT: This is a PORTRAIT image (taller than wide). Apply portrait-specific rules: Container at side, Border 5%, TT within 5% of border."
    else:
        asset_info += "\n\nORIENTATION: Unable to auto-detect (PDF input). Please determine orientation from the document content and apply the appropriate rules."

    # Add color mode instruction
    asset_info += f"\n\nCOLOR MODE: {color_mode}"
    if color_mode == "CMYK":
        asset_info += "\nIMPORTANT: This is an OOH or T-Sides asset. Check ALL colors using CMYK values (NOT RGB). Container must be CMYK 88, 45, 0, 0 for Prime Blue."
    else:
        asset_info += "\nIMPORTANT: This is a digital asset. Check ALL colors using RGB values (NOT CMYK). Container must be RGB 5, 120, 255 for Prime Blue."

    # Add Pre/Post CTA instruction
    asset_info += f"\n\nCTA TYPE: {pre_or_post.upper()}"
    if pre_or_post.lower() == "pre":
        asset_info += "\nIMPORTANT: This is a PRE-RELEASE asset. CTA MUST include a date (e.g., '21 JANUARY'). CTA MUST NOT have 'Watch Now' or similar action copy. Flag if CTA is missing a date."
    else:
        asset_info += "\nIMPORTANT: This is a POST-RELEASE asset. CTA MUST include action copy (e.g., 'WATCH NOW', 'NOW STREAMING'). CTA MUST NOT have a date. Flag if CTA includes a date."

    if asset_type and asset_type != "General":
        asset_info += f"\n\nAsset Type: {asset_type}\nIMPORTANT: Apply the specific compliance rules for this asset type (e.g., color mode requirements, logo placement rules, etc.)."

    if is_prime_original:
        asset_info += "\n\nContent Type: PRIME ORIGINAL\nIMPORTANT: This is Prime Original content. The 'Prime Original' logo MUST appear above the title treatment, and 'Original' must NOT appear in the CTA (to avoid redundancy)."
    else:
        asset_info += "\n\nContent Type: NON-PRIME ORIGINAL\nIMPORTANT: This is NOT Prime Original content. Only the 'Prime' logo should appear above the title treatment, and 'Original' MUST appear in the CTA."

    blocks.append({
        "type": "text",
        "text": asset_info,
    })

    # Add brand rules text if available
    if BRAND_RULES_PATH.exists():
        brand_rules = BRAND_RULES_PATH.read_text(encoding="utf-8")
        blocks.append({
            "type": "text",
            "text": f"=== BRAND COMPLIANCE RULES ===\n"
                    f"These are the MANDATORY rules that the image must follow. "
                    f"Check EVERY requirement meticulously and flag ANY deviations.\n\n"
                    f"{brand_rules}",
            "cache_control": {"type": "ephemeral"},
        })

    # Add guidelines if provided
    if guidelines_path:
        guide_data, guide_mime = encode_file(guidelines_path)
        if guide_mime == "application/pdf":
            blocks.append({
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": guide_mime,
                    "data": guide_data,
                },
                "cache_control": {"type": "ephemeral"},
            })
        else:
            blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": guide_mime,
                    "data": guide_data,
                },
                "cache_control": {"type": "ephemeral"},
            })
        blocks.append({
            "type": "text",
            "text": "Above is the brand guidelines document with visual examples. Use this as a reference to understand what correct layouts should look like, in combination with the explicit rules provided.",
        })

    # Add common mistakes examples if available
    if COMMON_MISTAKES_PATH.exists():
        mistakes_data, mistakes_mime = encode_file(COMMON_MISTAKES_PATH)
        blocks.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": mistakes_mime,
                "data": mistakes_data,
            },
            "cache_control": {"type": "ephemeral"},
        })
        blocks.append({
            "type": "text",
            "text": "=== COMMON MISTAKES EXAMPLES ===\n"
                    "Above are visual examples of common mistakes that must be detected and flagged. "
                    "Study these examples carefully to understand what kinds of errors to look for. "
                    "Each example shows a specific mistake that was NOT caught by previous checks. "
                    "Use these examples as training data to improve your detection accuracy.",
        })

    # Add master examples if available
    if MASTER_EXAMPLES_PATH.exists():
        master_data, master_mime = encode_file(MASTER_EXAMPLES_PATH)
        blocks.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": master_mime,
                "data": master_data,
            },
            "cache_control": {"type": "ephemeral"},
        })
        blocks.append({
            "type": "text",
            "text": f"=== MASTER EXAMPLES (PERFECT REFERENCES) ===\n"
                    f"Above are PERFECT, APPROVED examples for each asset type (OOH, DOOH, T-Sides, etc.). "
                    f"The top left of each page indicates the asset type.\n\n"
                    f"IMPORTANT: The image you are checking is a '{asset_type}' asset.\n"
                    f"Find the corresponding master example for '{asset_type}' in the document above.\n"
                    f"Compare the image being proofread against that specific master example.\n"
                    f"Check that measurements, positioning, colors, and layout match the master.\n"
                    f"Use the master as your reference for what 'correct' looks like for this asset type.",
        })

    # Add reference data from spreadsheets
    lang_list = [l.strip() for l in languages.split(",")]
    ref_data = load_reference_data(lang_list)
    if ref_data:
        blocks.append({
            "type": "text",
            "text": ref_data,
            # Note: Not cached due to 4-block limit. This data varies by language anyway.
        })

    blocks.append({
        "type": "text",
        "text": "Analyze the image and return the JSON report.",
    })

    return blocks


def format_report(data: dict) -> str:
    """Format the JSON response into a readable terminal report."""
    lines = []
    lines.append("=" * 60)
    lines.append("  PROOFREADER REPORT")
    lines.append("=" * 60)

    # Extracted text
    lines.append("\n--- Extracted Text ---")
    lines.append(data.get("extracted_text", "(none)"))

    # Languages
    langs = data.get("languages_detected", [])
    if langs:
        lines.append(f"\n--- Languages Detected ---")
        lines.append(", ".join(langs))

    # Element Identification
    element_id = data.get("element_identification", {})
    if element_id:
        lines.append(f"\n--- Element Identification ---")
        if element_id.get("title_treatment"):
            lines.append(f"Title Treatment (TT): {element_id['title_treatment']}")
        if element_id.get("signature"):
            lines.append(f"Signature:            {element_id['signature']}")
        if element_id.get("cta"):
            lines.append(f"CTA:                  {element_id['cta']}")
        if element_id.get("container_logo"):
            lines.append(f"Container Logo:       {element_id['container_logo']}")

    # Issues
    issues = data.get("issues", [])
    lines.append(f"\n--- Issues ({len(issues)} found) ---")

    if not issues:
        lines.append("No issues found.")
    else:
        severity_symbols = {"error": "[ERROR]", "warning": "[WARN] ", "info": "[INFO] "}
        for i, issue in enumerate(issues, 1):
            sev = severity_symbols.get(issue.get("severity", "info"), "[INFO] ")
            cat = issue.get("category", "unknown").upper()
            lines.append(f"\n  {i}. {sev} [{cat}]")
            if issue.get("location"):
                lines.append(f"     Location:   {issue['location']}")
            if issue.get("text"):
                lines.append(f"     Text:       {issue['text']}")
            lines.append(f"     Problem:    {issue.get('problem', 'N/A')}")
            lines.append(f"     Suggestion: {issue.get('suggestion', 'N/A')}")

    # Summary
    summary = data.get("summary", "")
    if summary:
        lines.append(f"\n--- Summary ---")
        lines.append(summary)

    lines.append("\n" + "=" * 60)
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Proofread images or PDFs for spelling, grammar, translation, and brand compliance issues."
    )
    parser.add_argument("image", type=Path, help="Path to the image (JPEG/PNG) or PDF to proofread")
    parser.add_argument("--guidelines", type=Path, default=None, help="Path to brand guidelines (PDF or image)")
    parser.add_argument("--languages", type=str, default="English", help='Expected languages, comma-separated (default: "English")')
    parser.add_argument("--asset-type", type=str, default="General",
                        choices=["General", "LRD", "OOH", "DOOH", "Digital Display", "Companion Banners", "T-Sides"],
                        help='Type of asset (default: "General")')
    parser.add_argument("--prime-original", action="store_true",
                        help="Flag if this is Prime Original content (affects logo and CTA requirements)")
    parser.add_argument("--pre-or-post", type=str, default="Pre",
                        choices=["Pre", "Post"],
                        help='CTA type: "Pre" (before release, must have date) or "Post" (after release, must have action copy like Watch Now)')
    parser.add_argument("--output", type=Path, default=None, help="Save report to a file")
    args = parser.parse_args()

    # Validate inputs
    if not args.image.exists():
        print(f"Error: Image file not found: {args.image}", file=sys.stderr)
        sys.exit(1)

    # Auto-load default brand guidelines if not explicitly specified
    if args.guidelines is None and BRAND_GUIDELINES_PATH.exists():
        args.guidelines = BRAND_GUIDELINES_PATH
        print(f"Auto-loading brand guidelines: {BRAND_GUIDELINES_PATH.name}", file=sys.stderr)

    if args.guidelines and not args.guidelines.exists():
        print(f"Error: Guidelines file not found: {args.guidelines}", file=sys.stderr)
        sys.exit(1)

    # Build and send API request
    client = anthropic.Anthropic()
    content = build_content_blocks(args.image, args.guidelines, args.languages, args.asset_type, args.prime_original, args.pre_or_post)

    print(f"Analyzing {args.image.name}...", file=sys.stderr)

    # Retry logic for API overload errors
    max_retries = 5
    retry_delay = 2  # seconds

    for attempt in range(max_retries):
        try:
            response = client.messages.create(
                model="claude-opus-4-6",
                max_tokens=4096,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": content}],
            )
            break  # Success, exit retry loop
        except anthropic.APIStatusError as e:
            if attempt < max_retries - 1:
                wait_time = retry_delay * (2 ** attempt)  # Exponential backoff
                print(f"⏳ API overloaded, retrying in {wait_time}s... (attempt {attempt + 1}/{max_retries})", file=sys.stderr)
                time.sleep(wait_time)
            else:
                print(f"❌ API overloaded after {max_retries} attempts. Please try again in a few minutes.", file=sys.stderr)
                sys.exit(1)

    # Extract text response
    raw_text = response.content[0].text

    # Parse JSON from response
    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError:
        # Try to extract JSON from markdown code blocks
        import re
        match = re.search(r"```(?:json)?\s*\n?(.*?)\n?```", raw_text, re.DOTALL)
        if match:
            data = json.loads(match.group(1))
        else:
            print("Error: Could not parse API response as JSON.", file=sys.stderr)
            print("Raw response:", file=sys.stderr)
            print(raw_text, file=sys.stderr)
            sys.exit(1)

    # Format and print report
    report = format_report(data)
    print(report)

    # Save to file if requested
    if args.output:
        args.output.write_text(report)
        print(f"\nReport saved to {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
